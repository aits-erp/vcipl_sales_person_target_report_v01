# # Copyright (c) 2026, Sukku and contributors
# # For license information, please see license.txt

# import io
# import frappe
# from frappe import _
# from frappe.utils import flt, getdate

# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
# from openpyxl.chart import BarChart, Reference
# from openpyxl.utils import get_column_letter


# MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
#           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# MONTH_FIELD_MAP = {
#     1: "jan_target", 2: "feb_target", 3: "mar_target",
#     4: "apr_target", 5: "may_target", 6: "jun_target",
#     7: "jul_target", 8: "aug_target", 9: "sep_target",
#     10: "oct_target", 11: "nov_target", 12: "dec_target"
# }

# DEFAULT_CATEGORIES = [
#     "Nonstick",
#     "Horeca",
#     "Pressure Cookers",
#     "SS Cookware",
#     "Healux",
#     "Kraft",
#     "Platinum",
#     "Platinum Triply P.cooker",
#     "Cast Iron",
#     "Bottle",
#     "Csd",
#     "Other",
#     "Futuretec"
# ]

# # --------------------------------------------------------------------
# # SECTION_VIEW_MAP
# #
# # Each entry: view_name -> (list_of_fields_to_check, list_of_values_to_match)
# #
# # A row belongs to the section if ANY of the listed fields, after
# # trimming whitespace and ignoring case, equals ANY of the listed
# # values (also trimmed/case-folded). Checking multiple fields matters
# # because a TSO can appear as either the "tso_name" on a row, OR as
# # the "parent_sales_person" (head) on someone else's row — depending
# # on how the Sales Person master data is set up. Mayur Talati, for
# # example, shows up as tso_name = "MAYUR TALATI" but his
# # parent_sales_person is "Sales Team", not himself — so we must match
# # on tso_name for him, while other heads are matched via
# # parent_sales_person.
# # --------------------------------------------------------------------
# SECTION_VIEW_MAP = {
#     "North Detail": (["parent_sales_person"], ["Rajiv K Dutta"]),
#     "East Detail": (["parent_sales_person"], ["Pannalal Bhattacharya"]),
#     "South Detail": (["parent_sales_person"], ["Mohammed Muqeemudheen Cherayakkuth"]),
#     # "West Detail": (["parent_sales_person", "tso_name"], [
#     #     "Sandeep Kumar",
#     #     "Jaydeo Deshmukh",
#     #     "Muslim Abdulla Hakim (Aslam)",
#     #     "MAYUR TALATI"
#     # ]),

#     "North TSO Detail": (["parent_sales_person"], ["Rajiv K Dutta"]),
#     "East TSO Detail": (["parent_sales_person"], ["Pannalal Bhattacharya"]),
#     "South TSO Detail": (["parent_sales_person"], ["Mohammed Muqeemudheen Cherayakkuth"]),
#     # "West TSO Detail": (["parent_sales_person", "tso_name"], [
#     #     "Sandeep Kumar",
#     #     "Jaydeo Deshmukh",
#     #     "Muslim Abdulla Hakim (Aslam)",
#     #     "MAYUR TALATI"
#     # ]),

#     "ROM Detail": (["parent_sales_person"], ["Sandeep Kumar"]),
#     "MPCG Detail": (["parent_sales_person", "tso_name"], ["Jaydeo Deshmukh"]),
#     "Mumbai AH Detail": (["parent_sales_person"], ["Muslim Abdulla Hakim (Aslam)"]),
#     "Gujarat Detail": (["tso_name", "parent_sales_person"], ["MAYUR TALATI"]),
# }

# _target_cache = {}


# def safe_field(category):
#     return category.replace(" ", "_").replace("-", "_")


# def normalize(value):
#     """Trim whitespace and case-fold so 'MAYUR TALATI', ' Mayur Talati ',
#     and 'mayur talati' are all treated as the same value."""
#     return (value or "").strip().casefold()


# def row_matches_section(row, fieldnames, selected_values):
#     """True if any of `fieldnames` on `row` normalized-equals any of
#     `selected_values`."""
#     normalized_targets = {normalize(v) for v in selected_values}

#     for fieldname in fieldnames:
#         if normalize(row.get(fieldname)) in normalized_targets:
#             return True

#     return False


# def filter_section_rows(rows, fieldnames, selected_values):
#     return [r for r in rows if row_matches_section(r, fieldnames, selected_values)]


# def execute(filters=None):
#     filters = frappe._dict(filters or {})

#     if filters.get("from_date") and filters.get("to_date"):
#         if getdate(filters.from_date) > getdate(filters.to_date):
#             frappe.throw(_("From Date must be before To Date"))

#     categories = get_categories(filters)
#     raw_data = get_data(filters, categories)

#     view_type = filters.get("view_type") or "TSO Summary"

#     if view_type == "Region Summary":
#         columns = get_region_summary_columns()
#         data = get_region_summary_data(raw_data)

#     elif view_type in SECTION_VIEW_MAP:
#         fieldnames, selected_values = SECTION_VIEW_MAP[view_type]
#         columns = get_columns(categories, filters)
#         data = filter_section_rows(raw_data, fieldnames, selected_values)

#         if view_type == "Gujarat Detail":
#             # Debug helper: if this ever comes back empty again, check
#             # the actual tso_name / parent_sales_person values being
#             # produced for Gujarat-related invoices.
#             distinct_tso = sorted({(r.get("tso_name") or "") for r in raw_data
#                                     if "mayur" in normalize(r.get("tso_name"))
#                                     or "mayur" in normalize(r.get("parent_sales_person"))})
#             frappe.errprint({"Gujarat Detail matched rows": len(data),
#                               "candidate tso_name/parent values seen": distinct_tso})

#     else:
#         columns = get_columns(categories, filters)
#         data = raw_data

#     summary = get_summary(data, categories)
#     chart = get_chart(data)

#     return columns, data, None, chart, summary


# def get_categories(filters):
#     if filters.get("custom_main_group"):
#         if isinstance(filters.custom_main_group, list):
#             return filters.custom_main_group
#         return [x.strip() for x in str(filters.custom_main_group).split(",") if x.strip()]

#     return DEFAULT_CATEGORIES


# def get_columns(categories, filters=None):
#     columns = []

#     if filters and filters.get("show_item_details"):
#         columns.extend([
#             {"label": _("Item Code"), "fieldname": "item_code", "fieldtype": "Link", "options": "Item", "width": 150},
#             {"label": _("Item Name"), "fieldname": "item_name", "fieldtype": "Data", "width": 200}
#         ])

#     columns.extend([
#         {"label": _("Month"), "fieldname": "month", "fieldtype": "Data", "width": 100},
#         {"label": _("TSO"), "fieldname": "tso_name", "fieldtype": "Data", "width": 200},
#         {"label": _("Customer Name"), "fieldname": "customer_name", "fieldtype": "Data", "width": 220},
#         {"label": _("Region"), "fieldname": "custom_region", "fieldtype": "Data", "width": 130},
#         {"label": _("Territory Code"), "fieldname": "custom_territory", "fieldtype": "Data", "width": 130},
#         {"label": _("Head Sales Person"), "fieldname": "parent_sales_person", "fieldtype": "Data", "width": 200},
#         {"label": _("Total Achieved"), "fieldname": "total_achieved", "fieldtype": "Currency", "width": 150},
#         {"label": _("Total Target"), "fieldname": "total_target", "fieldtype": "Currency", "width": 150},
#         {"label": _("Achievement %"), "fieldname": "achievement_percent", "fieldtype": "Percent", "width": 120},
#     ])

#     for cat in categories:
#         safe = safe_field(cat)
#         columns.append({"label": _(f"{cat} Target"), "fieldname": f"{safe}_target", "fieldtype": "Currency", "width": 130})
#         columns.append({"label": _(f"{cat} Achieved"), "fieldname": f"{safe}_achieved", "fieldtype": "Currency", "width": 130})

#     return columns


# def get_region_summary_columns():
#     return [
#         {"label": _("Region"), "fieldname": "region", "fieldtype": "Data", "width": 180},
#         {"label": _("Target"), "fieldname": "target", "fieldtype": "Currency", "width": 150},
#         {"label": _("Achieved"), "fieldname": "achieved", "fieldtype": "Currency", "width": 150},
#         {"label": _("Gap"), "fieldname": "gap", "fieldtype": "Currency", "width": 150},
#         {"label": _("Achievement %"), "fieldname": "achievement_percent", "fieldtype": "Percent", "width": 120},
#     ]


# def _load_targets_for_customer(customer, month_num):
#     cache_key = (customer, month_num)
#     if cache_key in _target_cache:
#         return _target_cache[cache_key]

#     month_field = MONTH_FIELD_MAP.get(int(month_num), "jan_target")

#     try:
#         rows = frappe.db.sql(f"""
#             SELECT
#                 mtd.main_group,
#                 mtd.{month_field} AS target_amount
#             FROM `tabMonthly Target Detail` mtd
#             INNER JOIN `tabSales Person Target` spt
#                 ON spt.name = mtd.parent
#             WHERE spt.name = %(customer)s
#               AND spt.period_type = 'Monthly'
#               AND spt.docstatus IN (0, 1)
#         """, {"customer": customer}, as_dict=1)

#         result = {r.main_group: flt(r.target_amount) for r in rows if r.main_group}
#         _target_cache[cache_key] = result
#         return result

#     except Exception as e:
#         frappe.log_error(f"Target fetch error for customer {customer}: {e}", "TEST REPORT")

#     _target_cache[cache_key] = {}
#     return {}


# def get_target_for_customer_category(customer, month_num, category):
#     targets = _load_targets_for_customer(customer, month_num)
#     return flt(targets.get(category, 0))


# def parse_multi(value):
#     if not value:
#         return []

#     if isinstance(value, list):
#         return value

#     return [x.strip() for x in str(value).split(",") if x.strip()]


# def get_data(filters, categories):
#     global _target_cache
#     _target_cache = {}

#     conditions = []
#     values = {}

#     if filters.get("from_date"):
#         conditions.append("si.posting_date >= %(from_date)s")
#         values["from_date"] = filters.get("from_date")

#     if filters.get("to_date"):
#         conditions.append("si.posting_date <= %(to_date)s")
#         values["to_date"] = filters.get("to_date")

#     if filters.get("sales_person"):
#         conditions.append("COALESCE(sp_inv.name, sp_cust.name) = %(sales_person)s")
#         values["sales_person"] = filters.get("sales_person")

#     if filters.get("parent_sales_person"):
#         conditions.append("""
#             COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person) = %(parent_sales_person)s
#         """)
#         values["parent_sales_person"] = filters.get("parent_sales_person")

#     if filters.get("customer"):
#         conditions.append("si.customer = %(customer)s")
#         values["customer"] = filters.get("customer")

#     if filters.get("customer_group"):
#         conditions.append("si.customer_group = %(customer_group)s")
#         values["customer_group"] = filters.get("customer_group")

#     regions = parse_multi(filters.get("custom_region"))
#     if regions:
#         conditions.append("COALESCE(sp_inv.custom_region, sp_cust.custom_region) IN %(custom_region)s")
#         values["custom_region"] = tuple(regions)

#     territories = parse_multi(filters.get("custom_territory"))
#     if territories:
#         conditions.append("COALESCE(sp_inv.custom_territory, sp_cust.custom_territory) IN %(custom_territory)s")
#         values["custom_territory"] = tuple(territories)

#     codes = parse_multi(filters.get("custom_head_sales_code"))
#     if codes:
#         conditions.append("COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code) IN %(custom_head_sales_code)s")
#         values["custom_head_sales_code"] = tuple(codes)

#     cat_filter = parse_multi(filters.get("custom_main_group"))
#     if cat_filter:
#         conditions.append("i.custom_main_group IN %(custom_main_group)s")
#         values["custom_main_group"] = tuple(cat_filter)

#     where_clause = " AND ".join(conditions) if conditions else "1=1"

#     group_by = """
#         DATE_FORMAT(si.posting_date, '%%Y-%%m'),
#         MONTH(si.posting_date),
#         YEAR(si.posting_date),
#         si.customer,
#         COALESCE(sp_inv.name, sp_cust.name, 'Unassigned'),
#         COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person, ''),
#         COALESCE(sp_inv.custom_region, sp_cust.custom_region, ''),
#         COALESCE(sp_inv.custom_territory, sp_cust.custom_territory, ''),
#         COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code, ''),
#         c.customer_name,
#         i.custom_main_group
#     """

#     if filters.get("show_item_details"):
#         group_by += ", sii.item_code, i.item_name"

#     query = f"""
#         SELECT
#             DATE_FORMAT(si.posting_date, '%%Y-%%m') AS month_key,
#             MONTH(si.posting_date) AS month_num,
#             YEAR(si.posting_date) AS year,
#             si.customer AS customer,
#             COALESCE(sp_inv.name, sp_cust.name, 'Unassigned') AS tso_name,
#             COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person, '') AS parent_sales_person,
#             COALESCE(sp_inv.custom_region, sp_cust.custom_region, '') AS custom_region,
#             COALESCE(sp_inv.custom_territory, sp_cust.custom_territory, '') AS custom_territory,
#             COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code, '') AS custom_head_sales_code,
#             c.customer_name,
#             i.custom_main_group AS category,
#             sii.item_code,
#             i.item_name,
#             SUM(sii.base_net_amount) AS achieved,
#             COUNT(DISTINCT si.name) AS invoice_count,
#             COUNT(DISTINCT sii.item_code) AS item_count
#         FROM `tabSales Invoice` si
#         INNER JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
#         INNER JOIN `tabItem` i ON i.name = sii.item_code
#         INNER JOIN `tabCustomer` c ON c.name = si.customer
#         LEFT JOIN `tabSales Team` st_inv ON st_inv.parent = si.name
#             AND st_inv.parenttype = 'Sales Invoice'
#             AND st_inv.idx = 1
#         LEFT JOIN `tabSales Person` sp_inv ON sp_inv.name = st_inv.sales_person
#         LEFT JOIN `tabSales Team` st_cust ON st_cust.parent = si.customer
#             AND st_cust.parenttype = 'Customer'
#         LEFT JOIN `tabSales Person` sp_cust ON sp_cust.name = st_cust.sales_person
#         WHERE si.docstatus = 1
#           AND i.custom_main_group IS NOT NULL
#           AND i.custom_main_group != ''
#           AND {where_clause}
#         GROUP BY {group_by}
#         ORDER BY YEAR(si.posting_date),
#                  MONTH(si.posting_date),
#                  custom_region,
#                  tso_name,
#                  c.customer_name
#     """

#     rows = frappe.db.sql(query, values, as_dict=1)
#     result = {}

#     for row in rows:
#         tso = row.tso_name or "Unassigned"
#         customer = row.customer

#         if filters.get("show_item_details"):
#             key = (row.month_key, customer, tso, row.item_code)
#         else:
#             key = (row.month_key, customer, tso)

#         if key not in result:
#             entry = {
#                 "month": f"{MONTHS[int(row.month_num)-1]}-{row.year}",
#                 "month_num": row.month_num,
#                 "year": row.year,
#                 "tso_name": tso,
#                 "customer_name": row.customer_name or "No Customer",
#                 "parent_sales_person": row.parent_sales_person or "",
#                 "custom_region": row.custom_region or "",
#                 "custom_territory": row.custom_territory or "",
#                 "custom_head_sales_code": row.custom_head_sales_code or "",
#                 "total_achieved": 0,
#                 "total_target": 0,
#                 "achievement_percent": 0,
#                 "invoice_count": 0,
#                 "item_count": 0
#             }

#             if filters.get("show_item_details"):
#                 entry["item_code"] = row.item_code
#                 entry["item_name"] = row.item_name

#             for cat in categories:
#                 safe = safe_field(cat)
#                 entry[f"{safe}_achieved"] = 0
#                 target_value = get_target_for_customer_category(customer, row.month_num, cat)
#                 entry[f"{safe}_target"] = flt(target_value)
#                 entry["total_target"] += flt(target_value)

#             result[key] = entry

#         if row.category in categories:
#             safe = safe_field(row.category)
#             result[key][f"{safe}_achieved"] += flt(row.achieved)

#         result[key]["total_achieved"] += flt(row.achieved)
#         result[key]["invoice_count"] += int(row.invoice_count or 0)
#         result[key]["item_count"] += int(row.item_count or 0)

#     final_data = list(result.values())

#     for row in final_data:
#         row["achievement_percent"] = (
#             flt(row.get("total_achieved")) / flt(row.get("total_target")) * 100
#             if flt(row.get("total_target")) else 0
#         )

#     return final_data


# def get_region_summary_data(data):
#     result = {}

#     for row in data:
#         region = row.get("custom_region") or "Unassigned"

#         if region not in result:
#             result[region] = {
#                 "region": region,
#                 "target": 0,
#                 "achieved": 0,
#                 "gap": 0,
#                 "achievement_percent": 0
#             }

#         result[region]["target"] += flt(row.get("total_target"))
#         result[region]["achieved"] += flt(row.get("total_achieved"))

#     final = []

#     for region, row in result.items():
#         row["gap"] = row["target"] - row["achieved"]
#         row["achievement_percent"] = (
#             row["achieved"] / row["target"] * 100
#             if row["target"] else 0
#         )
#         final.append(row)

#     return sorted(final, key=lambda x: x["region"])


# def get_summary(data, categories):
#     total_achieved = 0
#     total_target = 0
#     total_invoice = 0
#     total_item = 0

#     for row in data:
#         total_achieved += flt(row.get("total_achieved"))
#         total_target += flt(row.get("total_target"))
#         total_invoice += flt(row.get("invoice_count"))
#         total_item += flt(row.get("item_count"))

#     ach_percent = total_achieved / total_target * 100 if total_target else 0

#     return [
#         {"label": _("Total Achieved"), "value": total_achieved, "indicator": "Green", "datatype": "Currency"},
#         {"label": _("Total Target"), "value": total_target, "indicator": "Yellow", "datatype": "Currency"},
#         {"label": _("Achievement %"), "value": ach_percent, "indicator": "Purple", "datatype": "Percent"},
#         {"label": _("Invoice Count"), "value": total_invoice, "indicator": "Orange", "datatype": "Int"},
#         {"label": _("Item Count"), "value": total_item, "indicator": "Grey", "datatype": "Int"},
#     ]


# def get_chart(data):
#     region_data = get_region_summary_data(data)

#     if not region_data:
#         return None

#     return {
#         "data": {
#             "labels": [d["region"] for d in region_data],
#             "datasets": [
#                 {
#                     "name": "Target",
#                     "values": [d["target"] for d in region_data]
#                 },
#                 {
#                     "name": "Achieved",
#                     "values": [d["achieved"] for d in region_data]
#                 }
#             ]
#         },
#         "type": "bar",
#         "height": 280,
#     }


# @frappe.whitelist()
# def download_mis_excel(**filters):
#     filters = frappe._dict(filters or {})

#     categories = get_categories(filters)
#     data = get_data(filters, categories)

#     wb = Workbook()

#     ws = wb.active
#     ws.title = "TSO Summary"
#     make_tso_summary_sheet(ws, data, categories)

#     ws_region = wb.create_sheet("Region Summary")
#     make_region_summary_sheet(ws_region, data)

#     ws_chart = wb.create_sheet("Charts")
#     make_region_summary_sheet(ws_chart, data)
#     add_region_chart(ws_chart)

#     # Section-wise sheets like client workbook
#     for sheet_name, (fieldnames, selected_values) in SECTION_VIEW_MAP.items():
#         clean_sheet_name = sheet_name.replace(" Detail", "")[:31]

#         section_rows = filter_section_rows(data, fieldnames, selected_values)

#         if section_rows:
#             ws_sec = wb.create_sheet(clean_sheet_name)
#             make_tso_summary_sheet(ws_sec, section_rows, categories)

#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     frappe.local.response.filename = "Target_vs_Achievement_MIS.xlsx"
#     frappe.local.response.filecontent = output.read()
#     frappe.local.response.type = "download"


# def make_tso_summary_sheet(ws, data, categories):
#     dark_green = "0B6A57"
#     yellow = "FFF2CC"
#     blue = "DDEBF7"
#     total_fill = "D9EAD3"

#     thin = Side(style="thin", color="999999")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     headers = ["Region", "TSO", "Distributor"]

#     for cat in categories:
#         headers.append(f"{cat} Target")
#         headers.append(f"{cat} Achieved")

#     headers += ["Total Target", "Total Achieved", "Achievement %"]

#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
#     title = ws.cell(1, 1)
#     title.value = "Vinod Cookware — TSO Target vs Achievement"
#     title.font = Font(bold=True, color="FFFFFF", size=12)
#     title.fill = PatternFill("solid", fgColor=dark_green)
#     title.alignment = Alignment(horizontal="center")

#     for col, h in enumerate(headers, 1):
#         cell = ws.cell(2, col)
#         cell.value = h
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill("solid", fgColor=dark_green)
#         cell.border = border
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     row_no = 3

#     for row in data:
#         values = [
#             row.get("custom_region"),
#             row.get("tso_name"),
#             row.get("customer_name")
#         ]

#         for cat in categories:
#             safe = safe_field(cat)
#             values.append(flt(row.get(f"{safe}_target")))
#             values.append(flt(row.get(f"{safe}_achieved")))

#         total_target = flt(row.get("total_target"))
#         total_achieved = flt(row.get("total_achieved"))
#         ach_percent = total_achieved / total_target if total_target else 0

#         values += [total_target, total_achieved, ach_percent]

#         for col, value in enumerate(values, 1):
#             cell = ws.cell(row_no, col)
#             cell.value = value
#             cell.border = border
#             cell.alignment = Alignment(vertical="center", wrap_text=True)

#             if col > 3:
#                 cell.number_format = '#,##0.00'

#             if col >= 4 and col <= len(headers) - 3:
#                 if (col - 4) % 2 == 0:
#                     cell.fill = PatternFill("solid", fgColor=yellow)
#                 else:
#                     cell.fill = PatternFill("solid", fgColor=blue)

#             if col == len(headers):
#                 cell.number_format = "0.00%"

#         row_no += 1

#     total_row = row_no
#     ws.cell(total_row, 1).value = "Grand Total"
#     ws.cell(total_row, 1).font = Font(bold=True)

#     for col in range(4, len(headers)):
#         letter = get_column_letter(col)
#         cell = ws.cell(total_row, col)
#         cell.value = f"=SUM({letter}3:{letter}{total_row-1})"
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill("solid", fgColor=total_fill)
#         cell.border = border
#         cell.number_format = '#,##0.00'

#     ach_cell = ws.cell(total_row, len(headers))
#     total_target_col = get_column_letter(len(headers) - 2)
#     total_ach_col = get_column_letter(len(headers) - 1)
#     ach_cell.value = f"={total_ach_col}{total_row}/{total_target_col}{total_row}"
#     ach_cell.font = Font(bold=True)
#     ach_cell.fill = PatternFill("solid", fgColor=total_fill)
#     ach_cell.border = border
#     ach_cell.number_format = "0.00%"

#     ws.freeze_panes = "D3"
#     ws.auto_filter.ref = ws.dimensions

#     ws.column_dimensions["A"].width = 14
#     ws.column_dimensions["B"].width = 24
#     ws.column_dimensions["C"].width = 35

#     for col in range(4, len(headers) + 1):
#         ws.column_dimensions[get_column_letter(col)].width = 14


# def make_region_summary_sheet(ws, data):
#     dark_green = "0B6A57"
#     thin = Side(style="thin", color="999999")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     headers = ["Region", "Target", "Achieved", "Gap", "Achievement %"]

#     for col, h in enumerate(headers, 1):
#         cell = ws.cell(1, col)
#         cell.value = h
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill("solid", fgColor=dark_green)
#         cell.border = border
#         cell.alignment = Alignment(horizontal="center")

#     row_no = 2

#     for row in get_region_summary_data(data):
#         ws.cell(row_no, 1).value = row["region"]
#         ws.cell(row_no, 2).value = row["target"]
#         ws.cell(row_no, 3).value = row["achieved"]
#         ws.cell(row_no, 4).value = row["gap"]
#         ws.cell(row_no, 5).value = row["achievement_percent"] / 100

#         for col in range(1, 6):
#             c = ws.cell(row_no, col)
#             c.border = border
#             if col in [2, 3, 4]:
#                 c.number_format = '#,##0.00'
#             if col == 5:
#                 c.number_format = "0.00%"

#         row_no += 1

#     for col in range(1, 6):
#         ws.column_dimensions[get_column_letter(col)].width = 18


# def add_region_chart(ws):
#     if ws.max_row < 2:
#         return

#     chart = BarChart()
#     chart.title = "Region — Target vs Achieved"
#     chart.y_axis.title = "Amount"
#     chart.x_axis.title = "Region"

#     data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
#     cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

#     chart.add_data(data_ref, titles_from_data=True)
#     chart.set_categories(cats_ref)
#     chart.height = 10
#     chart.width = 20

#     ws.add_chart(chart, "G2")


# @frappe.whitelist()
# def get_mis_dashboard_data(from_date=None, to_date=None):
#     filters = frappe._dict({
#         "from_date": from_date,
#         "to_date": to_date,
#         "customer_group": "Debtors Distributors"
#     })

#     categories = get_categories(filters)
#     data = get_data(filters, categories)

#     # --------------------------
#     # KPI Summary
#     # --------------------------
#     total_target = sum(flt(d.get("total_target")) for d in data)
#     total_achieved = sum(flt(d.get("total_achieved")) for d in data)
#     total_gap = total_target - total_achieved
#     total_achievement = (
#         total_achieved / total_target * 100
#         if total_target else 0
#     )

#     # --------------------------
#     # Category Summary
#     # --------------------------
#     category_summary = []

#     for cat in categories:
#         safe = safe_field(cat)

#         target = sum(flt(d.get(f"{safe}_target")) for d in data)
#         achieved = sum(flt(d.get(f"{safe}_achieved")) for d in data)

#         gap = target - achieved

#         category_summary.append({
#             "category": cat,
#             "target": target,
#             "achieved": achieved,
#             "gap": gap,
#             "achievement": achieved / target * 100 if target else 0
#         })

#     # --------------------------
#     # Area Summary
#     #
#     # Same normalized, dual-field matching as SECTION_VIEW_MAP so the
#     # dashboard tiles stay consistent with the query report.
#     # --------------------------

#     AREA_FIELD_MAP = {
#         "ROM": (["parent_sales_person"], ["Sandeep Kumar"]),
#         "Mumbai AH": (["parent_sales_person"], ["Muslim Abdulla Hakim (Aslam)"]),
#         "Gujarat": (["tso_name", "parent_sales_person"], ["MAYUR TALATI"]),
#         "MPCG": (["parent_sales_person", "tso_name"], ["Jaydeo Deshmukh"]),
#     }

#     AREA_MAP = {
#         "North": lambda r: r.get("custom_region") == "North",
#         "South": lambda r: r.get("custom_region") == "South",
#         "East": lambda r: r.get("custom_region") == "East",
#         # "West": lambda r: r.get("custom_region") == "West",
#         "ROM": lambda r: row_matches_section(r, *AREA_FIELD_MAP["ROM"]),
#         "Mumbai AH": lambda r: row_matches_section(r, *AREA_FIELD_MAP["Mumbai AH"]),
#         "Gujarat": lambda r: row_matches_section(r, *AREA_FIELD_MAP["Gujarat"]),
#         "MPCG": lambda r: row_matches_section(r, *AREA_FIELD_MAP["MPCG"]),
#     }

#     area_summary = []

#     for area, condition in AREA_MAP.items():

#         rows = [x for x in data if condition(x)]

#         target = sum(flt(r.get("total_target")) for r in rows)
#         achieved = sum(flt(r.get("total_achieved")) for r in rows)

#         gap = target - achieved

#         area_summary.append({
#             "area": area,
#             "target": target,
#             "achieved": achieved,
#             "gap": gap,
#             "achievement": achieved / target * 100 if target else 0
#         })

#     return {
#         "kpi": {
#             "target": total_target,
#             "achieved": total_achieved,
#             "gap": total_gap,
#             "achievement": total_achievement
#         },
#         "category_summary": category_summary,
#         "area_summary": area_summary
#     }




# Copyright (c) 2026, Sukku and contributors
# For license information, please see license.txt

import io
import frappe
from frappe import _
from frappe.utils import flt, getdate

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

MONTH_FIELD_MAP = {
    1: "jan_target", 2: "feb_target", 3: "mar_target",
    4: "apr_target", 5: "may_target", 6: "jun_target",
    7: "jul_target", 8: "aug_target", 9: "sep_target",
    10: "oct_target", 11: "nov_target", 12: "dec_target"
}

DEFAULT_CATEGORIES = [
    "Nonstick",
    "Horeca",
    "Pressure Cookers",
    "SS Cookware",
    "Healux",
    "Kraft",
    "Platinum",
    "Platinum Triply P.cooker",
    "Cast Iron",
    "Bottle",
    "Csd",
    "Other",
    "Futuretec"
]

# --------------------------------------------------------------------
# SECTION_VIEW_MAP
#
# Each entry: view_name -> (list_of_fields_to_check, list_of_values_to_match)
#
# A row belongs to the section if ANY of the listed fields, after
# trimming whitespace and ignoring case, equals ANY of the listed
# values (also trimmed/case-folded). Checking multiple fields matters
# because a TSO can appear as either the "tso_name" on a row, OR as
# the "parent_sales_person" (head) on someone else's row — depending
# on how the Sales Person master data is set up. Mayur Talati, for
# example, shows up as tso_name = "MAYUR TALATI" but his
# parent_sales_person is "Sales Team", not himself — so we must match
# on tso_name for him, while other heads are matched via
# parent_sales_person.
#
# NOTE: "North Detail" / "North TSO Detail" (and the East/South
# equivalents) intentionally share the same match criteria. Both
# names are kept here because the query report's view_type filter and
# the dashboard's report-navigation buttons both rely on the
# "...TSO Detail" names existing. The Excel export below dedupes them
# back down to a single sheet per region — see get_dedup_section_map().
# --------------------------------------------------------------------
SECTION_VIEW_MAP = {
    "North Detail": (["parent_sales_person"], ["Rajiv K Dutta"]),
    "East Detail": (["parent_sales_person"], ["Pannalal Bhattacharya"]),
    "South Detail": (["parent_sales_person"], ["Mohammed Muqeemudheen Cherayakkuth"]),
    # "West Detail": (["parent_sales_person", "tso_name"], [
    #     "Sandeep Kumar",
    #     "Jaydeo Deshmukh",
    #     "Muslim Abdulla Hakim (Aslam)",
    #     "MAYUR TALATI"
    # ]),

    "North TSO Detail": (["parent_sales_person"], ["Rajiv K Dutta"]),
    "East TSO Detail": (["parent_sales_person"], ["Pannalal Bhattacharya"]),
    "South TSO Detail": (["parent_sales_person"], ["Mohammed Muqeemudheen Cherayakkuth"]),
    # "West TSO Detail": (["parent_sales_person", "tso_name"], [
    #     "Sandeep Kumar",
    #     "Jaydeo Deshmukh",
    #     "Muslim Abdulla Hakim (Aslam)",
    #     "MAYUR TALATI"
    # ]),

    "ROM Detail": (["parent_sales_person"], ["Sandeep Kumar"]),
    "MPCG Detail": (["parent_sales_person", "tso_name"], ["Jaydeo Deshmukh"]),
    "Mumbai AH Detail": (["parent_sales_person"], ["Muslim Abdulla Hakim (Aslam)"]),
    "Gujarat Detail": (["tso_name", "parent_sales_person"], ["MAYUR TALATI"]),
}

_target_cache = {}


def safe_field(category):
    return category.replace(" ", "_").replace("-", "_")


def normalize(value):
    """Trim whitespace and case-fold so 'MAYUR TALATI', ' Mayur Talati ',
    and 'mayur talati' are all treated as the same value."""
    return (value or "").strip().casefold()


def row_matches_section(row, fieldnames, selected_values):
    """True if any of `fieldnames` on `row` normalized-equals any of
    `selected_values`."""
    normalized_targets = {normalize(v) for v in selected_values}

    for fieldname in fieldnames:
        if normalize(row.get(fieldname)) in normalized_targets:
            return True

    return False


def filter_section_rows(rows, fieldnames, selected_values):
    return [r for r in rows if row_matches_section(r, fieldnames, selected_values)]


def get_dedup_section_map():
    """SECTION_VIEW_MAP intentionally has duplicate entries — e.g.
    "North Detail" and "North TSO Detail" both match the exact same
    rows (same fieldnames + same selected_values), because the report
    filter and the dashboard buttons both need the "...TSO Detail"
    names to exist. That's fine for filtering, but it means a naive
    loop over SECTION_VIEW_MAP would generate two identical sheets
    (e.g. "North" and "North TSO") in the Excel export.

    This collapses those duplicates down to a single entry per unique
    match criteria, keeping whichever name appears first in
    SECTION_VIEW_MAP (i.e. "North Detail" wins over "North TSO
    Detail"), so the workbook gets exactly one sheet per region.
    """
    seen_criteria = set()
    result = {}

    for sheet_name, (fieldnames, selected_values) in SECTION_VIEW_MAP.items():
        criteria_key = (tuple(fieldnames), tuple(selected_values))

        if criteria_key in seen_criteria:
            continue

        seen_criteria.add(criteria_key)
        result[sheet_name] = (fieldnames, selected_values)

    return result


def execute(filters=None):
    filters = frappe._dict(filters or {})

    if filters.get("from_date") and filters.get("to_date"):
        if getdate(filters.from_date) > getdate(filters.to_date):
            frappe.throw(_("From Date must be before To Date"))

    categories = get_categories(filters)
    raw_data = get_data(filters, categories)

    view_type = filters.get("view_type") or "TSO Summary"

    if view_type == "Region Summary":
        columns = get_region_summary_columns()
        data = get_region_summary_data(raw_data)

    elif view_type in SECTION_VIEW_MAP:
        fieldnames, selected_values = SECTION_VIEW_MAP[view_type]
        columns = get_columns(categories, filters)
        data = filter_section_rows(raw_data, fieldnames, selected_values)

        if view_type == "Gujarat Detail":
            # Debug helper: if this ever comes back empty again, check
            # the actual tso_name / parent_sales_person values being
            # produced for Gujarat-related invoices.
            distinct_tso = sorted({(r.get("tso_name") or "") for r in raw_data
                                    if "mayur" in normalize(r.get("tso_name"))
                                    or "mayur" in normalize(r.get("parent_sales_person"))})
            frappe.errprint({"Gujarat Detail matched rows": len(data),
                              "candidate tso_name/parent values seen": distinct_tso})

    else:
        columns = get_columns(categories, filters)
        data = raw_data

    summary = get_summary(data, categories)
    chart = get_chart(data)

    return columns, data, None, chart, summary


def get_categories(filters):
    if filters.get("custom_main_group"):
        if isinstance(filters.custom_main_group, list):
            return filters.custom_main_group
        return [x.strip() for x in str(filters.custom_main_group).split(",") if x.strip()]

    return DEFAULT_CATEGORIES


def get_columns(categories, filters=None):
    columns = []

    if filters and filters.get("show_item_details"):
        columns.extend([
            {"label": _("Item Code"), "fieldname": "item_code", "fieldtype": "Link", "options": "Item", "width": 150},
            {"label": _("Item Name"), "fieldname": "item_name", "fieldtype": "Data", "width": 200}
        ])

    columns.extend([
        {"label": _("Month"), "fieldname": "month", "fieldtype": "Data", "width": 100},
        {"label": _("TSO"), "fieldname": "tso_name", "fieldtype": "Data", "width": 200},
        {"label": _("Customer Name"), "fieldname": "customer_name", "fieldtype": "Data", "width": 220},
        {"label": _("Region"), "fieldname": "custom_region", "fieldtype": "Data", "width": 130},
        {"label": _("Territory Code"), "fieldname": "custom_territory", "fieldtype": "Data", "width": 130},
        {"label": _("Head Sales Person"), "fieldname": "parent_sales_person", "fieldtype": "Data", "width": 200},
        {"label": _("Total Achieved"), "fieldname": "total_achieved", "fieldtype": "Currency", "width": 150},
        {"label": _("Total Target"), "fieldname": "total_target", "fieldtype": "Currency", "width": 150},
        {"label": _("Achievement %"), "fieldname": "achievement_percent", "fieldtype": "Percent", "width": 120},
    ])

    for cat in categories:
        safe = safe_field(cat)
        columns.append({"label": _(f"{cat} Target"), "fieldname": f"{safe}_target", "fieldtype": "Currency", "width": 130})
        columns.append({"label": _(f"{cat} Achieved"), "fieldname": f"{safe}_achieved", "fieldtype": "Currency", "width": 130})

    return columns


def get_region_summary_columns():
    return [
        {"label": _("Region"), "fieldname": "region", "fieldtype": "Data", "width": 180},
        {"label": _("Target"), "fieldname": "target", "fieldtype": "Currency", "width": 150},
        {"label": _("Achieved"), "fieldname": "achieved", "fieldtype": "Currency", "width": 150},
        {"label": _("Gap"), "fieldname": "gap", "fieldtype": "Currency", "width": 150},
        {"label": _("Achievement %"), "fieldname": "achievement_percent", "fieldtype": "Percent", "width": 120},
    ]


def _load_targets_for_customer(customer, month_num):
    cache_key = (customer, month_num)
    if cache_key in _target_cache:
        return _target_cache[cache_key]

    month_field = MONTH_FIELD_MAP.get(int(month_num), "jan_target")

    try:
        rows = frappe.db.sql(f"""
            SELECT
                mtd.main_group,
                mtd.{month_field} AS target_amount
            FROM `tabMonthly Target Detail` mtd
            INNER JOIN `tabSales Person Target` spt
                ON spt.name = mtd.parent
            WHERE spt.name = %(customer)s
              AND spt.period_type = 'Monthly'
              AND spt.docstatus IN (0, 1)
        """, {"customer": customer}, as_dict=1)

        result = {r.main_group: flt(r.target_amount) for r in rows if r.main_group}
        _target_cache[cache_key] = result
        return result

    except Exception as e:
        frappe.log_error(f"Target fetch error for customer {customer}: {e}", "TEST REPORT")

    _target_cache[cache_key] = {}
    return {}


def get_target_for_customer_category(customer, month_num, category):
    targets = _load_targets_for_customer(customer, month_num)
    return flt(targets.get(category, 0))


def parse_multi(value):
    if not value:
        return []

    if isinstance(value, list):
        return value

    return [x.strip() for x in str(value).split(",") if x.strip()]


def get_data(filters, categories):
    global _target_cache
    _target_cache = {}

    conditions = []
    values = {}

    if filters.get("from_date"):
        conditions.append("si.posting_date >= %(from_date)s")
        values["from_date"] = filters.get("from_date")

    if filters.get("to_date"):
        conditions.append("si.posting_date <= %(to_date)s")
        values["to_date"] = filters.get("to_date")

    if filters.get("sales_person"):
        conditions.append("COALESCE(sp_inv.name, sp_cust.name) = %(sales_person)s")
        values["sales_person"] = filters.get("sales_person")

    if filters.get("parent_sales_person"):
        conditions.append("""
            COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person) = %(parent_sales_person)s
        """)
        values["parent_sales_person"] = filters.get("parent_sales_person")

    if filters.get("customer"):
        conditions.append("si.customer = %(customer)s")
        values["customer"] = filters.get("customer")

    if filters.get("customer_group"):
        conditions.append("si.customer_group = %(customer_group)s")
        values["customer_group"] = filters.get("customer_group")

    regions = parse_multi(filters.get("custom_region"))
    if regions:
        conditions.append("COALESCE(sp_inv.custom_region, sp_cust.custom_region) IN %(custom_region)s")
        values["custom_region"] = tuple(regions)

    territories = parse_multi(filters.get("custom_territory"))
    if territories:
        conditions.append("COALESCE(sp_inv.custom_territory, sp_cust.custom_territory) IN %(custom_territory)s")
        values["custom_territory"] = tuple(territories)

    codes = parse_multi(filters.get("custom_head_sales_code"))
    if codes:
        conditions.append("COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code) IN %(custom_head_sales_code)s")
        values["custom_head_sales_code"] = tuple(codes)

    cat_filter = parse_multi(filters.get("custom_main_group"))
    if cat_filter:
        conditions.append("i.custom_main_group IN %(custom_main_group)s")
        values["custom_main_group"] = tuple(cat_filter)

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    group_by = """
        DATE_FORMAT(si.posting_date, '%%Y-%%m'),
        MONTH(si.posting_date),
        YEAR(si.posting_date),
        si.customer,
        COALESCE(sp_inv.name, sp_cust.name, 'Unassigned'),
        COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person, ''),
        COALESCE(sp_inv.custom_region, sp_cust.custom_region, ''),
        COALESCE(sp_inv.custom_territory, sp_cust.custom_territory, ''),
        COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code, ''),
        c.customer_name,
        i.custom_main_group
    """

    if filters.get("show_item_details"):
        group_by += ", sii.item_code, i.item_name"

    query = f"""
        SELECT
            DATE_FORMAT(si.posting_date, '%%Y-%%m') AS month_key,
            MONTH(si.posting_date) AS month_num,
            YEAR(si.posting_date) AS year,
            si.customer AS customer,
            COALESCE(sp_inv.name, sp_cust.name, 'Unassigned') AS tso_name,
            COALESCE(sp_inv.parent_sales_person, sp_cust.parent_sales_person, '') AS parent_sales_person,
            COALESCE(sp_inv.custom_region, sp_cust.custom_region, '') AS custom_region,
            COALESCE(sp_inv.custom_territory, sp_cust.custom_territory, '') AS custom_territory,
            COALESCE(sp_inv.custom_head_sales_code, sp_cust.custom_head_sales_code, '') AS custom_head_sales_code,
            c.customer_name,
            i.custom_main_group AS category,
            sii.item_code,
            i.item_name,
            SUM(sii.base_net_amount) AS achieved,
            COUNT(DISTINCT si.name) AS invoice_count,
            COUNT(DISTINCT sii.item_code) AS item_count
        FROM `tabSales Invoice` si
        INNER JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        INNER JOIN `tabItem` i ON i.name = sii.item_code
        INNER JOIN `tabCustomer` c ON c.name = si.customer
        LEFT JOIN `tabSales Team` st_inv ON st_inv.parent = si.name
            AND st_inv.parenttype = 'Sales Invoice'
            AND st_inv.idx = 1
        LEFT JOIN `tabSales Person` sp_inv ON sp_inv.name = st_inv.sales_person
        LEFT JOIN `tabSales Team` st_cust ON st_cust.parent = si.customer
            AND st_cust.parenttype = 'Customer'
        LEFT JOIN `tabSales Person` sp_cust ON sp_cust.name = st_cust.sales_person
        WHERE si.docstatus = 1
          AND i.custom_main_group IS NOT NULL
          AND i.custom_main_group != ''
          AND {where_clause}
        GROUP BY {group_by}
        ORDER BY YEAR(si.posting_date),
                 MONTH(si.posting_date),
                 custom_region,
                 tso_name,
                 c.customer_name
    """

    rows = frappe.db.sql(query, values, as_dict=1)
    result = {}

    for row in rows:
        tso = row.tso_name or "Unassigned"
        customer = row.customer

        if filters.get("show_item_details"):
            key = (row.month_key, customer, tso, row.item_code)
        else:
            key = (row.month_key, customer, tso)

        if key not in result:
            entry = {
                "month": f"{MONTHS[int(row.month_num)-1]}-{row.year}",
                "month_num": row.month_num,
                "year": row.year,
                "tso_name": tso,
                "customer_name": row.customer_name or "No Customer",
                "parent_sales_person": row.parent_sales_person or "",
                "custom_region": row.custom_region or "",
                "custom_territory": row.custom_territory or "",
                "custom_head_sales_code": row.custom_head_sales_code or "",
                "total_achieved": 0,
                "total_target": 0,
                "achievement_percent": 0,
                "invoice_count": 0,
                "item_count": 0
            }

            if filters.get("show_item_details"):
                entry["item_code"] = row.item_code
                entry["item_name"] = row.item_name

            for cat in categories:
                safe = safe_field(cat)
                entry[f"{safe}_achieved"] = 0
                target_value = get_target_for_customer_category(customer, row.month_num, cat)
                entry[f"{safe}_target"] = flt(target_value)
                entry["total_target"] += flt(target_value)

            result[key] = entry

        if row.category in categories:
            safe = safe_field(row.category)
            result[key][f"{safe}_achieved"] += flt(row.achieved)

        result[key]["total_achieved"] += flt(row.achieved)
        result[key]["invoice_count"] += int(row.invoice_count or 0)
        result[key]["item_count"] += int(row.item_count or 0)

    final_data = list(result.values())

    for row in final_data:
        row["achievement_percent"] = (
            flt(row.get("total_achieved")) / flt(row.get("total_target")) * 100
            if flt(row.get("total_target")) else 0
        )

    return final_data


def get_region_summary_data(data):
    result = {}

    for row in data:
        region = row.get("custom_region") or "Unassigned"

        if region not in result:
            result[region] = {
                "region": region,
                "target": 0,
                "achieved": 0,
                "gap": 0,
                "achievement_percent": 0
            }

        result[region]["target"] += flt(row.get("total_target"))
        result[region]["achieved"] += flt(row.get("total_achieved"))

    final = []

    for region, row in result.items():
        row["gap"] = row["target"] - row["achieved"]
        row["achievement_percent"] = (
            row["achieved"] / row["target"] * 100
            if row["target"] else 0
        )
        final.append(row)

    return sorted(final, key=lambda x: x["region"])


def get_summary(data, categories):
    total_achieved = 0
    total_target = 0
    total_invoice = 0
    total_item = 0

    for row in data:
        total_achieved += flt(row.get("total_achieved"))
        total_target += flt(row.get("total_target"))
        total_invoice += flt(row.get("invoice_count"))
        total_item += flt(row.get("item_count"))

    ach_percent = total_achieved / total_target * 100 if total_target else 0

    return [
        {"label": _("Total Achieved"), "value": total_achieved, "indicator": "Green", "datatype": "Currency"},
        {"label": _("Total Target"), "value": total_target, "indicator": "Yellow", "datatype": "Currency"},
        {"label": _("Achievement %"), "value": ach_percent, "indicator": "Purple", "datatype": "Percent"},
        {"label": _("Invoice Count"), "value": total_invoice, "indicator": "Orange", "datatype": "Int"},
        {"label": _("Item Count"), "value": total_item, "indicator": "Grey", "datatype": "Int"},
    ]


def get_chart(data):
    region_data = get_region_summary_data(data)

    if not region_data:
        return None

    return {
        "data": {
            "labels": [d["region"] for d in region_data],
            "datasets": [
                {
                    "name": "Target",
                    "values": [d["target"] for d in region_data]
                },
                {
                    "name": "Achieved",
                    "values": [d["achieved"] for d in region_data]
                }
            ]
        },
        "type": "bar",
        "height": 280,
    }


@frappe.whitelist()
def download_mis_excel(**filters):
    filters = frappe._dict(filters or {})

    categories = get_categories(filters)
    data = get_data(filters, categories)

    wb = Workbook()

    ws = wb.active
    ws.title = "TSO Summary"
    make_tso_summary_sheet(ws, data, categories)

    ws_region = wb.create_sheet("Region Summary")
    make_region_summary_sheet(ws_region, data)

    ws_chart = wb.create_sheet("Charts")
    make_region_summary_sheet(ws_chart, data)
    add_region_chart(ws_chart)

    # Section-wise sheets like client workbook.
    # Uses the deduped map so regions that are matched under two names
    # (e.g. "North Detail" / "North TSO Detail") only produce ONE sheet.
    for sheet_name, (fieldnames, selected_values) in get_dedup_section_map().items():
        clean_sheet_name = sheet_name.replace(" Detail", "")[:31]

        section_rows = filter_section_rows(data, fieldnames, selected_values)

        if section_rows:
            ws_sec = wb.create_sheet(clean_sheet_name)
            make_tso_summary_sheet(ws_sec, section_rows, categories)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = "Target_vs_Achievement_MIS.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "download"


def make_tso_summary_sheet(ws, data, categories):
    dark_green = "0B6A57"
    yellow = "FFF2CC"
    blue = "DDEBF7"
    total_fill = "D9EAD3"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Region", "TSO", "Distributor"]

    for cat in categories:
        headers.append(f"{cat} Target")
        headers.append(f"{cat} Achieved")

    headers += ["Total Target", "Total Achieved", "Achievement %"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(1, 1)
    title.value = "Vinod Cookware — TSO Target vs Achievement"
    title.font = Font(bold=True, color="FFFFFF", size=12)
    title.fill = PatternFill("solid", fgColor=dark_green)
    title.alignment = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark_green)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_no = 3

    for row in data:
        values = [
            row.get("custom_region"),
            row.get("tso_name"),
            row.get("customer_name")
        ]

        for cat in categories:
            safe = safe_field(cat)
            values.append(flt(row.get(f"{safe}_target")))
            values.append(flt(row.get(f"{safe}_achieved")))

        total_target = flt(row.get("total_target"))
        total_achieved = flt(row.get("total_achieved"))
        ach_percent = total_achieved / total_target if total_target else 0

        values += [total_target, total_achieved, ach_percent]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row_no, col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col > 3:
                cell.number_format = '#,##0.00'

            if col >= 4 and col <= len(headers) - 3:
                if (col - 4) % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=yellow)
                else:
                    cell.fill = PatternFill("solid", fgColor=blue)

            if col == len(headers):
                cell.number_format = "0.00%"

        row_no += 1

    total_row = row_no
    ws.cell(total_row, 1).value = "Grand Total"
    ws.cell(total_row, 1).font = Font(bold=True)

    for col in range(4, len(headers)):
        letter = get_column_letter(col)
        cell = ws.cell(total_row, col)
        cell.value = f"=SUM({letter}3:{letter}{total_row-1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=total_fill)
        cell.border = border
        cell.number_format = '#,##0.00'

    ach_cell = ws.cell(total_row, len(headers))
    total_target_col = get_column_letter(len(headers) - 2)
    total_ach_col = get_column_letter(len(headers) - 1)
    ach_cell.value = f"={total_ach_col}{total_row}/{total_target_col}{total_row}"
    ach_cell.font = Font(bold=True)
    ach_cell.fill = PatternFill("solid", fgColor=total_fill)
    ach_cell.border = border
    ach_cell.number_format = "0.00%"

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 35

    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def make_region_summary_sheet(ws, data):
    dark_green = "0B6A57"
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Region", "Target", "Achieved", "Gap", "Achievement %"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark_green)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row_no = 2

    for row in get_region_summary_data(data):
        ws.cell(row_no, 1).value = row["region"]
        ws.cell(row_no, 2).value = row["target"]
        ws.cell(row_no, 3).value = row["achieved"]
        ws.cell(row_no, 4).value = row["gap"]
        ws.cell(row_no, 5).value = row["achievement_percent"] / 100

        for col in range(1, 6):
            c = ws.cell(row_no, col)
            c.border = border
            if col in [2, 3, 4]:
                c.number_format = '#,##0.00'
            if col == 5:
                c.number_format = "0.00%"

        row_no += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18


def add_region_chart(ws):
    if ws.max_row < 2:
        return

    chart = BarChart()
    chart.title = "Region — Target vs Achieved"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Region"

    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 20

    ws.add_chart(chart, "G2")


@frappe.whitelist()
def get_mis_dashboard_data(from_date=None, to_date=None):
    filters = frappe._dict({
        "from_date": from_date,
        "to_date": to_date,
        "customer_group": "Debtors Distributors"
    })

    categories = get_categories(filters)
    data = get_data(filters, categories)

    # --------------------------
    # KPI Summary
    # --------------------------
    total_target = sum(flt(d.get("total_target")) for d in data)
    total_achieved = sum(flt(d.get("total_achieved")) for d in data)
    total_gap = total_target - total_achieved
    total_achievement = (
        total_achieved / total_target * 100
        if total_target else 0
    )

    # --------------------------
    # Category Summary
    # --------------------------
    category_summary = []

    for cat in categories:
        safe = safe_field(cat)

        target = sum(flt(d.get(f"{safe}_target")) for d in data)
        achieved = sum(flt(d.get(f"{safe}_achieved")) for d in data)

        gap = target - achieved

        category_summary.append({
            "category": cat,
            "target": target,
            "achieved": achieved,
            "gap": gap,
            "achievement": achieved / target * 100 if target else 0
        })

    # --------------------------
    # Area Summary
    #
    # Same normalized, dual-field matching as SECTION_VIEW_MAP so the
    # dashboard tiles stay consistent with the query report.
    # --------------------------

    AREA_FIELD_MAP = {
        "ROM": (["parent_sales_person"], ["Sandeep Kumar"]),
        "Mumbai AH": (["parent_sales_person"], ["Muslim Abdulla Hakim (Aslam)"]),
        "Gujarat": (["tso_name", "parent_sales_person"], ["MAYUR TALATI"]),
        "MPCG": (["parent_sales_person", "tso_name"], ["Jaydeo Deshmukh"]),
    }

    AREA_MAP = {
        "North": lambda r: r.get("custom_region") == "North",
        "South": lambda r: r.get("custom_region") == "South",
        "East": lambda r: r.get("custom_region") == "East",
        # "West": lambda r: r.get("custom_region") == "West",
        "ROM": lambda r: row_matches_section(r, *AREA_FIELD_MAP["ROM"]),
        "Mumbai AH": lambda r: row_matches_section(r, *AREA_FIELD_MAP["Mumbai AH"]),
        "Gujarat": lambda r: row_matches_section(r, *AREA_FIELD_MAP["Gujarat"]),
        "MPCG": lambda r: row_matches_section(r, *AREA_FIELD_MAP["MPCG"]),
    }

    area_summary = []

    for area, condition in AREA_MAP.items():

        rows = [x for x in data if condition(x)]

        target = sum(flt(r.get("total_target")) for r in rows)
        achieved = sum(flt(r.get("total_achieved")) for r in rows)

        gap = target - achieved

        area_summary.append({
            "area": area,
            "target": target,
            "achieved": achieved,
            "gap": gap,
            "achievement": achieved / target * 100 if target else 0
        })

    return {
        "kpi": {
            "target": total_target,
            "achieved": total_achieved,
            "gap": total_gap,
            "achievement": total_achievement
        },
        "category_summary": category_summary,
        "area_summary": area_summary
    }